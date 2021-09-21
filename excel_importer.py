import unicodedata
from os import listdir, path
from os.path import isfile, join

import openpyxl

import variables


class ExcelImporter:

    @classmethod
    def get_data(cls):
        for original_file_name in listdir(variables.EXCEL_FOLDER):
            if not isfile(join(variables.EXCEL_FOLDER, original_file_name)):
                continue
            if original_file_name[0] in ['.', '~', '$']:
                continue
            original_file_name = unicodedata.normalize('NFC', original_file_name)
            file_name, ext = path.splitext(original_file_name)
            if not ext or ext not in ('.xlsx', '.xls'):
                continue
            print(f'START IMPORT FILE {original_file_name}')
            datas = cls.import_excel(variables.EXCEL_FOLDER + original_file_name)
            print(f'END IMPORT FILE {original_file_name}')
            return datas

    @classmethod
    def import_excel(cls, file):
        excel_file = openpyxl.load_workbook(file)
        datas = {}
        for sheet in excel_file.worksheets:
            sheet_name = sheet.title
            coords = []
            for row_idx in range(2, sheet.max_row + 1):
                coord = (
                    sheet.cell(row=row_idx, column=1).value,
                    sheet.cell(row=row_idx, column=2).value,
                    sheet.cell(row=row_idx, column=3).value
                )
                coords.append(coord)
            datas[sheet_name] = coords
        return datas